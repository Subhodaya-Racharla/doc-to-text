import io
import re
import time
from difflib import SequenceMatcher
from typing import List

import fitz  # PyMuPDF
import openpyxl
import pytesseract
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from PIL import Image

app = FastAPI(title="doc-to-text API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "https://frontend-tau-rose-13.vercel.app"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tiff", ".bmp"}
OCR_TEXT_THRESHOLD = 50  # characters; below this a page is treated as scanned


def file_extension(filename: str) -> str:
    dot = filename.rfind(".")
    return filename[dot:].lower() if dot != -1 else ""


def ocr_page(page: fitz.Page) -> str:
    """Render a PDF page to a PIL Image and run Tesseract OCR on it."""
    pixmap = page.get_pixmap(dpi=300)
    image = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
    return pytesseract.image_to_string(image)


@app.get("/health")
async def health_check():
    return {"status": "ok"}


@app.post("/extract-text")
async def extract_text(file: UploadFile = File(...)):
    if file_extension(file.filename) not in PDF_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only PDF files are supported.")

    contents = await file.read()

    try:
        doc = fitz.open(stream=contents, filetype="pdf")
    except Exception:
        raise HTTPException(
            status_code=422,
            detail="Could not open file. It may be corrupted or not a valid PDF.",
        )

    pages = []
    start = time.perf_counter()
    try:
        for i in range(len(doc)):
            page = doc.load_page(i)
            text = page.get_text()

            if len(text.strip()) < OCR_TEXT_THRESHOLD:
                try:
                    text = ocr_page(page)
                    method = "ocr"
                except pytesseract.TesseractNotFoundError:
                    raise HTTPException(
                        status_code=500,
                        detail="Tesseract is not installed or not found in PATH. Please install it on the server.",
                    )
                except Exception as e:
                    raise HTTPException(status_code=500, detail=f"OCR failed on page {i + 1}: {str(e)}")
            else:
                method = "direct"

            pages.append({
                "page_number": i + 1,
                "text": text,
                "method": method,
                "character_count": len(text),
            })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading page: {str(e)}")
    finally:
        doc.close()

    extraction_time = round(time.perf_counter() - start, 2)

    return {
        "filename": file.filename,
        "total_pages": len(pages),
        "total_characters": sum(p["character_count"] for p in pages),
        "extraction_time_seconds": extraction_time,
        "pages": pages,
    }


@app.post("/extract-text-image")
async def extract_text_image(file: UploadFile = File(...)):
    if file_extension(file.filename) not in IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Accepted formats: PNG, JPG, JPEG, TIFF, BMP.",
        )

    contents = await file.read()

    try:
        image = Image.open(io.BytesIO(contents))
        image.verify()  # confirm it's a valid image
    except Exception:
        raise HTTPException(
            status_code=422,
            detail="Could not open image. The file may be corrupted or is not a valid image.",
        )

    try:
        # Re-open after verify() — verify() exhausts the stream
        image = Image.open(io.BytesIO(contents))
        text = pytesseract.image_to_string(image)
    except pytesseract.TesseractNotFoundError:
        raise HTTPException(
            status_code=500,
            detail="Tesseract is not installed or not found in PATH. Please install it on the server.",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR failed: {str(e)}")

    return {
        "filename": file.filename,
        "text": text,
    }


def extract_full_text(contents: bytes) -> str:
    """Extract all text from a PDF, using OCR for scanned pages."""
    doc = fitz.open(stream=contents, filetype="pdf")
    all_text = []
    try:
        for i in range(len(doc)):
            page = doc.load_page(i)
            text = page.get_text()
            if len(text.strip()) < OCR_TEXT_THRESHOLD:
                try:
                    text = ocr_page(page)
                except Exception:
                    pass
            all_text.append(text)
    finally:
        doc.close()
    return "\n".join(all_text)


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 1: OCR Text Cleaning
# ═══════════════════════════════════════════════════════════════════════════════

_NOISE_PATTERNS = [
    re.compile(r"(?i)^\s*(tap to zoom|powered by|page \d+ of \d+|confidential|"
               r"scanned by|printed on|generated by|all rights reserved|"
               r"www\.\S+|http\S+|©.*\d{4})\s*$"),
    re.compile(r"^\s*[-_=*]{4,}\s*$"),
    re.compile(r"^\s*\d+\s*$"),
    re.compile(r"^\s*.{0,2}\s*$"),
]


def clean_ocr_text(text: str) -> str:
    """Clean OCR output: remove noise, merge broken lines, normalize whitespace."""
    lines = text.split("\n")
    cleaned = []
    for line in lines:
        if any(p.match(line) for p in _NOISE_PATTERNS):
            continue
        line = re.sub(r"[ \t]+", " ", line).rstrip()
        cleaned.append(line)

    merged = []
    for line in cleaned:
        stripped = line.strip()
        if not stripped:
            merged.append("")
            continue
        if (merged and merged[-1] and
                not merged[-1].rstrip().endswith((":", ".", ",", ";", "-", "=")) and
                stripped and stripped[0].islower()):
            merged[-1] = merged[-1].rstrip() + " " + stripped
        else:
            merged.append(line)

    return "\n".join(merged)


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 2: Field Normalization, Similarity Helpers & Alias Mapping
# ═══════════════════════════════════════════════════════════════════════════════

def _normalize(s: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    s = s.lower()
    s = re.sub(r"[.,:;!?()'\"\-/#]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _strip_all_punct(s: str) -> str:
    """Remove ALL non-alphanumeric characters and lowercase."""
    return re.sub(r"[^a-zA-Z0-9]", "", s).lower()


def _char_similarity(a: str, b: str) -> float:
    """Character-level similarity (0.0 to 1.0) using SequenceMatcher."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


_ALIASES: dict[str, list[str]] = {
    "name": ["full name", "applicant name", "candidate name", "patient name",
             "insured name", "member name", "employee name", "claimant name",
             "beneficiary name", "subscriber name", "holder name"],
    "dob": ["date of birth", "d o b", "birth date", "birthdate", "born on",
            "date of birth dob"],
    "gender": ["sex", "m f"],
    "age": ["age yrs", "age years"],
    "address": ["present address", "current address", "residential address",
                "mailing address", "permanent address", "home address",
                "street address", "correspondence address"],
    "city": ["town", "place"],
    "state": ["province"],
    "zip": ["zip code", "zipcode", "postal code", "pin", "pin code", "pincode"],
    "license no": ["licence no", "license number", "licence number",
                    "dl no", "dl number", "driving license no",
                    "driving licence no", "license", "licence"],
    "id no": ["id number", "identification no", "identification number",
              "id", "govt id"],
    "ssn": ["social security number", "social security no", "ss no", "ss number"],
    "passport no": ["passport number", "passport"],
    "phone": ["phone no", "phone number", "mobile", "mobile no", "mobile number",
              "contact no", "contact number", "tel", "telephone",
              "cell", "cell phone"],
    "email": ["email id", "email address", "e mail", "e-mail", "mail id"],
    "claim number": ["claim no", "claim id", "claim #", "claim ref",
                     "reference number", "ref no", "reference no"],
    "policy number": ["policy no", "policy id", "policy #",
                      "member id", "member no", "subscriber id"],
    "group number": ["group no", "group id", "group #", "grp no"],
    "dos": ["date of service", "service date", "dates of service"],
    "diagnosis": ["dx", "diagnosis code", "icd", "icd code", "icd 10"],
    "cpt": ["cpt code", "procedure code", "proc code", "hcpcs"],
    "denial reason": ["reason for denial", "denial code", "denial",
                      "reason denied", "denied reason"],
    "provider": ["provider name", "rendering provider", "doctor",
                 "physician", "attending", "referring provider"],
    "patient id": ["patient no", "patient number", "mrn",
                   "medical record number", "medical record no"],
    "npi": ["npi number", "npi no", "national provider identifier"],
    "amount": ["total amount", "amount due", "total due", "balance",
               "amount billed", "billed amount", "charged amount",
               "total charges", "net amount", "grand total", "total"],
    "invoice no": ["invoice number", "invoice #", "inv no", "bill no",
                   "bill number", "receipt no", "receipt number"],
    "date": ["invoice date", "bill date", "statement date", "issue date",
             "effective date"],
    "due date": ["payment due", "due by", "payable by"],
    "tax": ["tax amount", "gst", "vat", "sales tax"],
    "discount": ["discount amount", "rebate"],
    "employer": ["employer name", "company", "company name", "organization"],
    "occupation": ["designation", "job title", "position", "role"],
    "salary": ["pay", "wage", "compensation", "ctc", "annual salary"],
    # Driving license specific
    "s/w/d": ["s w d", "son of", "wife of", "daughter of", "s o", "w o", "d o",
              "father name", "father s name", "husband name"],
    "blood group": ["blood grp", "blood type", "bl group", "bl grp"],
    "date of issue": ["issue date", "issued on", "doi", "dt of issue",
                      "date of issue"],
    "date of expiry": ["expiry date", "valid till", "valid until", "doe",
                       "expires on", "exp date", "validity", "date of expiry"],
    "authorization to drive": ["auth to drive", "authorized to drive",
                                "vehicle class", "class of vehicle", "cov",
                                "category", "categories", "authorization to drive"],
}


def _get_all_variants(field: str) -> set[str]:
    """Get all normalized name variants for a field, including aliases."""
    norm = _normalize(field)
    variants = {norm}

    for key, aliases in _ALIASES.items():
        nk = _normalize(key)
        all_forms = [nk] + [_normalize(a) for a in aliases]
        if norm in all_forms:
            variants.update(all_forms)

    for key, aliases in _ALIASES.items():
        nk = _normalize(key)
        all_forms = [nk] + [_normalize(a) for a in aliases]
        for form in all_forms:
            if len(norm) >= 3 and (norm in form or form in norm):
                variants.update(all_forms)
                break

    return variants


def _field_variants(field: str) -> list[str]:
    """Generate regex patterns for all variants of a user-supplied field name."""
    variants = _get_all_variants(field)
    result = []

    for v in sorted(variants, key=len, reverse=True):
        words = v.split()
        parts = []
        for w in words:
            ocr_word = re.escape(w)
            ocr_word = ocr_word.replace("c", "[ce]").replace("e", "[ec]")
            ocr_word = ocr_word.replace("i", "[il1]").replace("l", "[li1]")
            ocr_word = ocr_word.replace("o", "[o0]").replace("0", "[0o]")
            ocr_word = ocr_word.replace("s", "[s5]").replace("5", "[5s]")
            parts.append(ocr_word + r"[.,:;#]?")
        # Normal variant with flexible separators between words
        pattern = r"[\s\-_]*".join(parts)
        result.append(pattern)

        # Also add a "joined" variant for OCR word-joining errors
        # e.g., "authorization to" also matches "authorizationto"
        if len(words) > 1:
            joined = "".join(words)
            ocr_joined = re.escape(joined)
            ocr_joined = ocr_joined.replace("c", "[ce]").replace("e", "[ec]")
            ocr_joined = ocr_joined.replace("i", "[il1]").replace("l", "[li1]")
            ocr_joined = ocr_joined.replace("o", "[o0]").replace("0", "[0o]")
            ocr_joined = ocr_joined.replace("s", "[s5]").replace("5", "[5s]")
            result.append(ocr_joined + r"[.,:;#]?")

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 3: Field Position Detection (regex + fuzzy similarity)
# ═══════════════════════════════════════════════════════════════════════════════

def _find_field_position(text: str, field: str) -> tuple[int, int, str] | None:
    """Find where a field label appears in text.

    Returns (start_pos, value_start_pos, matched_keyword) or None.
    Tries regex patterns first, then falls back to fuzzy similarity scanning.
    """
    variants = _field_variants(field)

    # Strategy 1: LABEL followed by separator (:, =, -, tab)
    for vp in variants:
        m = re.search(rf"(?im)({vp})\s*[:=\-\u2014\t]\s*", text)
        if m:
            return m.start(), m.end(), m.group(1).strip()

    # Strategy 2: LABEL followed by wide whitespace
    for vp in variants:
        m = re.search(rf"(?im)^[ \t]*({vp})\s{{2,}}", text)
        if m:
            return m.start(), m.end(), m.group(1).strip()

    # Strategy 3: LABEL at end of line
    for vp in variants:
        m = re.search(rf"(?im)({vp})\s*$", text)
        if m:
            return m.start(), m.end(), m.group(1).strip()

    # Strategy 4: LABEL in text followed by separator or newline
    for vp in variants:
        m = re.search(rf"(?i)({vp})", text)
        if m:
            after = text[m.end():m.end() + 10]
            if re.match(r"\s*[:=\-\u2014\t]", after) or re.match(r"\s*\n", after):
                return m.start(), m.end(), m.group(1).strip()

    # Strategy 5: Fuzzy similarity scan (handles OCR errors like Agaress→Address)
    return _fuzzy_find_field(text, field)


def _fuzzy_find_field(text: str, field: str) -> tuple[int, int, str] | None:
    """Scan lines for labels with >=70% character similarity to the field name."""
    all_variants = _get_all_variants(field)

    lines = text.split("\n")
    best_score = 0.0
    best_result: tuple[int, int, str] | None = None

    offset = 0
    for line in lines:
        stripped = line.strip()
        if not stripped:
            offset += len(line) + 1
            continue

        label = ""
        value_start_in_line = -1

        # Try colon separator first
        colon_pos = stripped.find(":")
        if colon_pos > 0:
            label = stripped[:colon_pos].strip()
            # Find position of colon in original line
            colon_in_line = line.find(":")
            value_start_in_line = colon_in_line + 1
            # Skip whitespace after colon
            while value_start_in_line < len(line) and line[value_start_in_line] in " \t":
                value_start_in_line += 1
        else:
            # Try equals separator
            eq_pos = stripped.find("=")
            if eq_pos > 0:
                label = stripped[:eq_pos].strip()
                eq_in_line = line.find("=")
                value_start_in_line = eq_in_line + 1
                while value_start_in_line < len(line) and line[value_start_in_line] in " \t":
                    value_start_in_line += 1
            else:
                # Try wide space separator
                sep_m = re.match(r"^(.*?)\s{2,}(.+)$", stripped)
                if sep_m:
                    label = sep_m.group(1).strip()
                    value_start_in_line = line.find(sep_m.group(2))

        if not label or len(label) > 80 or value_start_in_line < 0:
            offset += len(line) + 1
            continue

        label_norm = _normalize(label)
        label_nospace = _strip_all_punct(label)

        for variant in all_variants:
            variant_nospace = _strip_all_punct(variant)

            sim = max(
                _char_similarity(label_norm, variant),
                _char_similarity(label_nospace, variant_nospace),
            )

            if sim > best_score and sim >= 0.7:
                best_score = sim
                best_result = (offset, offset + value_start_in_line, label)

        offset += len(line) + 1

    return best_result


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 4: Value Cleanup
# ═══════════════════════════════════════════════════════════════════════════════

def _clean_value(value: str) -> str:
    """Clean extracted value: strip leading separators, extra whitespace, etc."""
    # Remove leading colons, dashes, equals, spaces
    value = re.sub(r"^[\s:=\-\u2014]+", "", value)
    # Remove trailing separators
    value = re.sub(r"[\s:=\-\u2014]+$", "", value)
    # Collapse internal whitespace
    value = re.sub(r"\s+", " ", value).strip()
    return value


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 5: Multi-line & Full-Block Value Extraction
# ═══════════════════════════════════════════════════════════════════════════════

def _is_kv_line(line: str) -> bool:
    """Does this line look like a LABEL : VALUE pattern?"""
    stripped = line.strip()
    if not stripped:
        return False
    return bool(re.match(r"^[A-Za-z][A-Za-z\s/.#()&,]{0,60}?\s*:\s*\S", stripped))


def _is_field_label(line: str, all_field_patterns: list[str] | None = None) -> bool:
    """Heuristic: does this line look like the start of a new key-value field?"""
    stripped = line.strip()
    if not stripped:
        return False

    if _is_kv_line(line):
        return True
    if re.match(r"^[A-Z][A-Z\s./#]{1,50}\s*[:=\-]\s*", stripped):
        return True
    if stripped.count("|") >= 2:
        return True
    if all_field_patterns:
        for fp in all_field_patterns:
            if re.match(rf"(?i)^\s*{fp}\s*[:=\-\s]", stripped):
                return True

    return False


def _extract_multiline_value(text: str, start_pos: int,
                             all_field_patterns: list[str] | None = None,
                             max_lines: int = 20) -> str:
    """Extract value starting at start_pos, spanning multiple lines until next field."""
    after = text[start_pos:]
    lines = after.split("\n")
    if not lines:
        return ""

    # Skip initial empty lines (up to 2)
    start_idx = 0
    while start_idx < len(lines) and not lines[start_idx].strip():
        start_idx += 1
        if start_idx > 2:
            return ""

    value_parts = []
    consecutive_empty = 0
    for i in range(start_idx, min(len(lines), start_idx + max_lines)):
        stripped = lines[i].strip()

        if not stripped:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                break
            if value_parts:
                found_more = False
                for j in range(i + 1, min(len(lines), i + 3)):
                    ns = lines[j].strip()
                    if ns:
                        if _is_field_label(ns, all_field_patterns):
                            break
                        found_more = True
                        break
                if not found_more:
                    break
            continue

        consecutive_empty = 0

        # Stop if this line starts a new KV field (only after first value line)
        if value_parts and _is_field_label(stripped, all_field_patterns):
            break

        value_parts.append(stripped)

    # Join multi-line values with comma+space
    value = ", ".join(value_parts) if len(value_parts) > 1 else " ".join(value_parts)
    return _clean_value(value)


def _extract_full_block(text: str, field: str,
                        all_field_patterns: list[str] | None = None) -> tuple[str, str]:
    """Full Block extraction. Returns (value, matched_keyword)."""
    result = _find_field_position(text, field)
    if not result:
        return "", ""

    _start, end, matched = result
    after = text[end:]

    next_field_pos = len(after)

    kv_pattern = re.compile(
        r"^[ \t]*[A-Za-z][A-Za-z\s./#]{1,50}\s*[:=]\s*",
        re.MULTILINE
    )
    for m in kv_pattern.finditer(after):
        if m.start() > 0:
            next_field_pos = min(next_field_pos, m.start())
            break

    if all_field_patterns:
        for fp in all_field_patterns:
            pm = re.search(rf"(?im)^[ \t]*{fp}\s*[:=\-\s]", after)
            if pm and pm.start() > 0:
                next_field_pos = min(next_field_pos, pm.start())

    block = after[:next_field_pos].strip()
    lines = [l.strip() for l in block.split("\n") if l.strip()]
    value = ", ".join(lines) if len(lines) > 1 else " ".join(lines)
    return _clean_value(value), matched


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 6: Table Detection & Extraction
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_table_rows(text: str) -> list[dict[str, str]]:
    """Detect and extract tabular data from text (pipe or tab delimited)."""
    lines = text.split("\n")
    tables: list[dict[str, str]] = []

    pipe_lines = [(i, line) for i, line in enumerate(lines)
                  if line.strip().count("|") >= 2]
    if len(pipe_lines) >= 2:
        groups: list[list[tuple[int, str]]] = []
        current: list[tuple[int, str]] = [pipe_lines[0]]
        for j in range(1, len(pipe_lines)):
            if pipe_lines[j][0] - pipe_lines[j - 1][0] <= 2:
                current.append(pipe_lines[j])
            else:
                if len(current) >= 2:
                    groups.append(current)
                current = [pipe_lines[j]]
        if len(current) >= 2:
            groups.append(current)

        for group in groups:
            rows_text = [g[1] for g in group]
            header_cells = [c.strip() for c in rows_text[0].split("|") if c.strip()]
            for row_text in rows_text[1:]:
                if re.match(r"^[\s|+\-=]+$", row_text):
                    continue
                cells = [c.strip() for c in row_text.split("|") if c.strip()]
                if len(cells) >= len(header_cells):
                    row_dict = {}
                    for hi, header in enumerate(header_cells):
                        row_dict[header] = cells[hi] if hi < len(cells) else ""
                    tables.append(row_dict)

    if not tables:
        tab_lines = [(i, line) for i, line in enumerate(lines)
                     if line.count("\t") >= 1 and len(line.strip()) > 3]
        if len(tab_lines) >= 2:
            header_cells = [c.strip() for c in tab_lines[0][1].split("\t") if c.strip()]
            if len(header_cells) >= 2:
                for _, row_text in tab_lines[1:]:
                    cells = [c.strip() for c in row_text.split("\t")]
                    if cells:
                        row_dict = {}
                        for hi, header in enumerate(header_cells):
                            row_dict[header] = cells[hi] if hi < len(cells) else ""
                        tables.append(row_dict)

    return tables


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 7: Auto-Detect Fields
# ═══════════════════════════════════════════════════════════════════════════════

def auto_detect_fields(text: str) -> list[dict[str, str]]:
    """Scan text for key-value patterns and return detected fields with preview values."""
    detected: list[dict[str, str]] = []
    seen_labels: set[str] = set()

    # Pattern: "Label : Value" or "Label = Value" or "LABEL - Value"
    patterns = [
        re.compile(r"^[ \t]*([A-Za-z][A-Za-z\s./#]{1,50}?)\s*[:=]\s*(.+)$", re.MULTILINE),
        re.compile(r"^[ \t]*([A-Z][A-Z\s./#]{1,50}?)\s*[-\u2014]\s*(.+)$", re.MULTILINE),
    ]

    for pattern in patterns:
        for m in pattern.finditer(text):
            label = m.group(1).strip()
            value = m.group(2).strip()

            # Skip if label is too short or too generic
            if len(label) < 2 or len(label) > 50:
                continue
            # Skip if value is empty or looks like a separator
            if not value or re.match(r"^[-_=.]+$", value):
                continue
            # Skip duplicate labels (case-insensitive)
            norm_label = label.lower().strip()
            if norm_label in seen_labels:
                continue
            seen_labels.add(norm_label)

            # Truncate preview value
            preview = value[:80] + ("\u2026" if len(value) > 80 else "")
            detected.append({"field": label, "preview": preview})

    # Limit to reasonable number
    return detected[:30]


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 7.5: KV Format Detection & Extraction
# ═══════════════════════════════════════════════════════════════════════════════

def _is_kv_document(text: str) -> bool:
    """Check if document primarily uses LABEL : VALUE format."""
    lines = [l for l in text.split("\n") if l.strip()]
    if len(lines) < 3:
        return False
    kv_count = sum(1 for l in lines if _is_kv_line(l))
    return kv_count >= 3 and kv_count / len(lines) >= 0.2


def _extract_kv_value(text: str, field: str) -> tuple[str, str]:
    """Extract value from a LABEL : VALUE format document using fuzzy matching.

    Returns (extracted_value, matched_label_in_document).
    """
    all_variants = _get_all_variants(field)

    lines = text.split("\n")
    best_score = 0.0
    best_label = ""
    best_line_idx = -1

    for i, line in enumerate(lines):
        colon_pos = line.find(":")
        if colon_pos < 1:
            continue
        label = line[:colon_pos].strip()
        if not label:
            continue

        label_norm = _normalize(label)
        label_nospace = _strip_all_punct(label)

        for variant in all_variants:
            variant_nospace = _strip_all_punct(variant)

            # Exact match gets score 1.0
            if label_norm == variant:
                score = 1.0
            else:
                score = max(
                    _char_similarity(label_norm, variant),
                    _char_similarity(label_nospace, variant_nospace),
                )

            if score > best_score:
                best_score = score
                best_label = label
                best_line_idx = i

    if best_score < 0.7 or best_line_idx < 0:
        return "", ""

    # Extract value after the colon
    line = lines[best_line_idx]
    colon_pos = line.find(":")
    value_start = line[colon_pos + 1:].strip()
    value_lines = [value_start] if value_start else []

    # Continue to next lines until we hit another KV line
    for j in range(best_line_idx + 1, min(len(lines), best_line_idx + 20)):
        next_line = lines[j].strip()
        if not next_line:
            if value_lines:
                break
            continue
        if _is_kv_line(lines[j]):
            break
        value_lines.append(next_line)

    value = ", ".join(value_lines) if len(value_lines) > 1 else " ".join(value_lines)
    return _clean_value(value), best_label


# ═══════════════════════════════════════════════════════════════════════════════
# MODULE 8: Main Field Extraction (combines all modules)
# ═══════════════════════════════════════════════════════════════════════════════

def extract_field_value(text: str, field: str,
                        all_field_patterns: list[str] | None = None,
                        method: str = "auto") -> tuple[str, str]:
    """Extract a field value from text.

    Returns (extracted_value, matched_keyword_in_document).

    Methods:
        auto   - Try KV format, then after-keyword, then full-block, then table
        block  - Full block extraction (everything until next field)
    """
    # Strategy 0: KV document format (most reliable for structured docs)
    if _is_kv_document(text):
        value, matched = _extract_kv_value(text, field)
        if value:
            return value, matched

    # Strategy 1: Block method (if requested)
    if method == "block":
        value, matched = _extract_full_block(text, field, all_field_patterns)
        if value:
            return value, matched

    # Strategy 2: After-keyword extraction with multi-line support
    result = _find_field_position(text, field)
    if result:
        start, end, matched = result
        value = _extract_multiline_value(text, end, all_field_patterns)
        if value:
            return value, matched
        # If after-keyword found nothing, try full block
        value, matched2 = _extract_full_block(text, field, all_field_patterns)
        if value:
            return value, matched2 or matched

    # Strategy 3: Fallback to table lookup
    tables = _extract_table_rows(text)
    norm_field = _normalize(field)
    for table_row in tables:
        for col_name, col_val in table_row.items():
            if _normalize(col_name) == norm_field and col_val.strip():
                return _clean_value(col_val), col_name

    return "", ""


async def _read_file_text(file: UploadFile) -> str:
    """Read and extract text from a single uploaded file."""
    ext = file_extension(file.filename)
    contents = await file.read()

    if ext in PDF_EXTENSIONS:
        text = extract_full_text(contents)
    elif ext in IMAGE_EXTENSIONS:
        try:
            image = Image.open(io.BytesIO(contents))
            text = pytesseract.image_to_string(image)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"OCR failed for {file.filename}: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file: {file.filename}")

    return clean_ocr_text(text)


async def _extract_rows(files: List[UploadFile], fields: str,
                         method: str = "auto") -> tuple[list[str], list[dict], list[dict], list[dict]]:
    """Extract fields from files.

    Returns (field_list, rows, context_rows, matched_keyword_rows).
    """
    import json

    try:
        field_list = json.loads(fields)
    except (json.JSONDecodeError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid fields format. Send a JSON array of strings.")

    if not field_list or not isinstance(field_list, list):
        raise HTTPException(status_code=400, detail="Provide at least one field name.")

    rows = []
    context_rows = []
    matched_keyword_rows = []
    for file in files:
        text = await _read_file_text(file)

        all_patterns = []
        for f in field_list:
            all_patterns.extend(_field_variants(f))

        row = {"Filename": file.filename}
        ctx = {"Filename": file.filename}
        kw = {"Filename": file.filename}
        for field in field_list:
            value, matched_keyword = extract_field_value(text, field, all_patterns, method)
            row[field] = value
            kw[field] = matched_keyword if matched_keyword else ""

            # Build context: show where the match was found
            if value:
                pos = _find_field_position(text, field)
                if pos:
                    pos_start, pos_end, _ = pos
                    start = max(0, pos_start - 10)
                    val_pos = text.find(value.split(",")[0].strip(), pos_end)
                    if val_pos >= 0:
                        end = min(len(text), val_pos + len(value) + 20)
                    else:
                        end = min(len(text), pos_end + len(value) + 40)
                    snippet = text[start:end].replace("\n", " \u21b5 ").strip()
                    ctx[field] = snippet
                else:
                    ctx[field] = ""
            else:
                ctx[field] = ""
        rows.append(row)
        context_rows.append(ctx)
        matched_keyword_rows.append(kw)

    return field_list, rows, context_rows, matched_keyword_rows


@app.post("/extract-fields")
async def extract_fields(
    files: List[UploadFile] = File(...),
    fields: str = Form(...),
    method: str = Form("auto"),
):
    """Extract user-defined fields from multiple PDFs and return JSON preview."""
    field_list, rows, context_rows, matched_keyword_rows = await _extract_rows(files, fields, method)
    return {
        "fields": field_list,
        "rows": rows,
        "context": context_rows,
        "matched_keywords": matched_keyword_rows,
    }


@app.post("/auto-detect-fields")
async def auto_detect_fields_endpoint(
    file: UploadFile = File(...),
):
    """Scan a document and auto-detect extractable fields."""
    text = await _read_file_text(file)
    detected = auto_detect_fields(text)
    return {"detected_fields": detected}


@app.post("/extract-fields-xlsx")
async def extract_fields_xlsx(
    files: List[UploadFile] = File(...),
    fields: str = Form(...),
):
    """Extract user-defined fields and return an Excel spreadsheet."""
    from openpyxl.styles import Alignment, Font, PatternFill

    field_list, rows, _, _ = await _extract_rows(files, fields)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    headers = ["Filename"] + field_list
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=extracted_data.xlsx"},
    )
